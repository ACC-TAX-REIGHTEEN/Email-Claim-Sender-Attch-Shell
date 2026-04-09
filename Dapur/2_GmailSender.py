import smtplib
import ssl
import configparser
import os
import openpyxl
from email.message import EmailMessage

print("--> Membaca file konfigurasi config.conf")
config = configparser.ConfigParser()
config.read('config.conf')

sender_email = config['SMTP']['sender_email']
sender_password = config['SMTP']['sender_password']
to_email = config['RECIPIENT']['to_email']
cc_emails = config['RECIPIENT']['cc_email']
body_template = config['CONTENT']['body']

print("--> Membaca data dari isian_temp.xlsx")
try:
    wb = openpyxl.load_workbook('isian_temp.xlsx', data_only=True)
    sheet = wb['Isian']
    
    header_row = None
    col_invoice = None
    col_program = None
    
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        for col_idx, cell_value in enumerate(row):
            if cell_value == "No Invoice Klaim":
                col_invoice = col_idx
                header_row = row_idx
            elif cell_value == "Nama Program Klaim":
                col_program = col_idx
        if header_row is not None and col_program is not None:
            break
            
    if header_row is not None and col_invoice is not None and col_program is not None:
        context = ssl.create_default_context()
        
        print("--> Membuka koneksi SMTP Gmail")
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
            server.login(sender_email, sender_password)
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
                no_invoice = row[col_invoice]
                nama_program = row[col_program]
                
                if no_invoice and nama_program:
                    no_invoice = str(no_invoice).strip()
                    nama_program = str(nama_program).strip()
                    
                    subject = "INV PT. ABCXYZ " + no_invoice
                    
                    pdf_filename = no_invoice.replace('-', '').replace('/', '') + ".pdf"
                    
                    if os.path.exists(pdf_filename):
                        msg = EmailMessage()
                        msg['Subject'] = subject
                        msg['From'] = sender_email
                        msg['To'] = to_email
                        msg['Cc'] = cc_emails
                        
                        body = body_template.replace('{nama_program}', nama_program)
                        msg.set_content(body, subtype='html')
                        
                        with open(pdf_filename, 'rb') as f:
                            pdf_data = f.read()
                            msg.add_attachment(pdf_data, maintype='application', subtype='pdf', filename=pdf_filename)
                        
                        server.send_message(msg)
                        print("--> Email untuk " + no_invoice + " berhasil dikirim beserta lampiran " + pdf_filename)
                    else:
                        print("--> File PDF " + pdf_filename + " untuk invoice " + no_invoice + " tidak ditemukan")
                        
        print("--> Seluruh proses pengiriman selesai")
    else:
        print("--> Kolom No Invoice Klaim atau Nama Program Klaim tidak ditemukan")
        
except FileNotFoundError:
    print("--> File isian_temp.xlsx tidak ditemukan")
except Exception as e:
    print("--> Terjadi kesalahan: " + str(e))