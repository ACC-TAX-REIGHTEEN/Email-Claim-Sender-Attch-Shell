import openpyxl
from openpyxl.utils import get_column_letter

source_file = "Surat Saffiela - 130326.xlsm"
target_file = "isian_temp.xlsx"
sheet_name = "Isian"

print("--> Memulai proses ekstraksi file")

try:
    wb_source = openpyxl.load_workbook(source_file, data_only=True)
    
    if sheet_name in wb_source.sheetnames:
        sheet_source = wb_source[sheet_name]
        
        wb_target = openpyxl.Workbook()
        sheet_target = wb_target.active
        sheet_target.title = sheet_name
        
        print("--> Menyalin data ke sheet baru")
        
        for row in sheet_source.iter_rows(values_only=True):
            sheet_target.append(row)
            
        print("--> Melakukan auto-fit pada kolom")
        
        for col in sheet_target.columns:
            max_length = 0
            column_letter = col[0].column_letter
            
            for cell in col:
                if cell.value is not None:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                        
            adjusted_width = max_length + 2
            sheet_target.column_dimensions[column_letter].width = adjusted_width
            
        wb_target.save(target_file)
        print("--> File isian_temp.xlsx berhasil dibuat dan disimpan")
        
    else:
        print("--> Sheet Isian tidak ditemukan di dalam file sumber")
        
except FileNotFoundError:
    print("--> File Surat Saffiela - 130326.xlsm tidak ditemukan")
except Exception as e:
    print("--> Terjadi kesalahan: " + str(e))